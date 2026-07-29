from time import time

from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import pandas as pd
from api.utils import private_supabase
from api.views import get_ppmp_items
from rest_framework.response import Response

def is_empty_or_zero(value):
    return pd.isna(value) or value == 0

def testingPPMP(excel_file, row_start, name_column, unit_column, quantity_column, price_per_unit_column, year, ppmp_category="Office Supply"):
    fiscal_year_str = year
    fiscal_year = private_supabase.table("FISCAL_YEAR").select("*").eq("Year", fiscal_year_str).execute()


    df = pd.read_excel(excel_file, header=None, skiprows=row_start - 1)

    required_columns = [
        name_column,
        unit_column,
        quantity_column,
        price_per_unit_column,
    ]

    missing = [c for c in required_columns if c not in df.columns]

    if missing:
        raise ValueError({
            "message": f"Column(s) {missing} do not exist or are completely empty.",
        })

    current_category = None
    processed_rows = []

    for _, row in df.iterrows():
        description = row[name_column]
        unit = row[unit_column]
        quantity = row[quantity_column]
        price = row[price_per_unit_column]
        # print(row[name_column-1], description, unit, quantity, price)
        left = row[name_column - 1]
        right = row[name_column]

        name = None

        if pd.notna(left):
            name = str(left).strip()
        elif pd.notna(right):
            name = str(right).strip()
        else:
            continue

        if(
            pd.notna(row[name_column])
            and is_empty_or_zero(unit)
            and is_empty_or_zero(quantity)
            and is_empty_or_zero(price)
        ): # check if category
            if "subtotal" in name.lower() or "total" in name.lower():
                continue
            current_category = name
            continue
        elif(
            pd.notna(row[name_column-1])
            and is_empty_or_zero(unit)
            and is_empty_or_zero(quantity)
            and is_empty_or_zero(price)
        ):
            if "subtotal" in name.lower() or "total" in name.lower():
                continue
            current_category = name
            continue
        if (
            pd.notna(description)
            and pd.notna(unit)
            and pd.notna(quantity)
            and pd.notna(price)
        ): # legit
            processed_rows.append({
                "Description": description,
                "Unit": unit,
                "Quantity": quantity,
                "CatalogPrice": f"{price:.2f}",
                "Category": current_category
            })

    df = pd.DataFrame(processed_rows)

    # check for incorrect data types (mga NaN)
    quantity = pd.to_numeric(df["Quantity"], errors="coerce")
    price = pd.to_numeric(df["CatalogPrice"], errors="coerce")

    bad = df[quantity.isna() | price.isna()]

    if not bad.empty:
        errors = []

        for index, row in bad.iterrows():
            errors.append({
                "row": index,
                "quantity": row["Quantity"],
                "price": row["CatalogPrice"],
            })
        raise ValueError({
            "message": "Invalid numeric values found in the Excel file.",
            "rows": errors,
        })

    df["Quantity"] = quantity
    df["CatalogPrice"] = price

    df["TotalAmount"] = df["Quantity"] * df["CatalogPrice"]
    total_amount =  df["TotalAmount"].sum()
    if fiscal_year.data:
        return df, total_amount, True
    else:
        return df, total_amount, False

def upload_excel(df, total_ABC, year, ppmp_category="Office Supply"):
    fiscal_year = private_supabase.table("FISCAL_YEAR").select("*").eq("Year", year).execute()
    fiscal_year_id = 0
    if fiscal_year.data and ppmp_category == "Office Supply":
        fiscal_year_id = fiscal_year.data[0]["FiscalYearID"]
        private_supabase.table("FISCAL_YEAR").delete().eq("FiscalYearID", fiscal_year_id).execute() #cascade delete

        response = private_supabase.table("FISCAL_YEAR").insert({
                "Year": year,
                "TotalABC": total_ABC,
                "Status": "ongoing"
            }).execute()
        fiscal_year_id = response.data[0]["FiscalYearID"]
    else:
        response = private_supabase.table("FISCAL_YEAR").select("FiscalYearID").eq("Year", year).maybe_single().execute()
        if response is None:
            response = private_supabase.table("FISCAL_YEAR").insert({
                "Year": year,
                "TotalABC": total_ABC,
                "Status": "ongoing"
            }).execute()
            fiscal_year_id = response.data[0]["FiscalYearID"]
        else:
            fiscal_year_id = response.data["FiscalYearID"]
    records = []

    for _, row in df.iterrows():
        records.append({
            "ItemName": row["Description"],
            "UnitName": row["Unit"],
            "PlannedQuantity": int(row["Quantity"]),
            "AvailableQuantity": int(row["Quantity"]),
            "PricePerUnit": float(row["CatalogPrice"]),
            "PendingQuantity": 0,
            "FulfilledQuantity": 0,
            "FiscalYearID": fiscal_year_id,
            "ItemCategory": row["Category"],
            "PpmpCategory": ppmp_category,
        })
    try:
        private_supabase.table("PPMP_ITEM").insert(records).execute()
    except TypeError as e:
        return e

def export_formatted_excel(year):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    title = "CICT-PPMP-" + year
    ws.title = title
    default_font = Font(name="Arial", size=10)
    total_count = 0
    grand_total_amount = 0
    start_row = 0
    end_row = 0
    start_column = 1
    end_column = 0
    start_number_column = 0
    end_number_column = 0
    start_decimal_column = 0
    end_decimal_column = 0

    current_row = 2
    current_column = 1
    (
        current_column,
        current_row,
        start_number_column,
        end_number_column,
        start_decimal_column,
        end_decimal_column,
        end_column,
    ) = set_header(ws, current_column, current_row, year)
    ppmp_items = get_ppmp_items(year)

    office_supplies = [ppmp_item for ppmp_item in ppmp_items.data if ppmp_item["PpmpCategory"] == "Office Supply"]
    lab_supplies = [ppmp_item for ppmp_item in ppmp_items.data if ppmp_item["PpmpCategory"] == "Laboratory Supply/Equipment"]

    office_categories = [ppmp_item["ItemCategory"] for ppmp_item in office_supplies]
    office_categories = list(dict.fromkeys(office_categories))
    lab_categories = [ppmp_item["ItemCategory"] for ppmp_item in lab_supplies]
    lab_categories = list(dict.fromkeys(lab_categories))
    office_supplies_dict = {}

    for office_category in office_categories:
        category_items = [
            item
            for item in office_supplies
            if item["ItemCategory"] == office_category
        ]

        office_supplies_dict[office_category] = [
            {
                "Seq.": i,
                "GENERAL DESCRIPTION": item["ItemName"],
                "Unit of Measure": item["UnitName"],
                "January": item["PlannedQuantity"],
                "TOTAL": item["PlannedQuantity"],
                "Price as per Catalogue": item["PricePerUnit"],
                "TOTAL AMOUNT": item["PlannedQuantity"] * item["PricePerUnit"],
            }
            for i, item in enumerate(category_items, start=1)
        ]

    office_supplies = office_supplies_dict

    lab_supplies_dict = {}

    for lab_category in lab_categories:
        category_items = [
            item
            for item in lab_supplies
            if item["ItemCategory"] == lab_category
        ]

        lab_supplies_dict[lab_category] = [
            {
                "Seq.": i,
                "GENERAL DESCRIPTION": item["ItemName"],
                "Unit of Measure": item["UnitName"],
                "January": item["PlannedQuantity"],
                "TOTAL": item["PlannedQuantity"],
                "Price as per Catalogue": item["PricePerUnit"],
                "TOTAL AMOUNT": item["PlannedQuantity"] * item["PricePerUnit"],
            }
            for i, item in enumerate(category_items, start=1)
        ]

    lab_supplies = lab_supplies_dict


    # return Response({"Office": office_supplies, "Lab": lab_supplies})
    # display subtotal per category
    # sa subtotal price per cat = 0.00
    # get signatories

    current_row += 2
    current_column = 1

    ws[f"{num_to_letter(current_column)}{current_row}"] = "OFFICE SUPPLIES"
    ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column + 1)}{current_row}")
    set_border_to_cell(ws, current_column, current_row, left=None, right=None, top=None, bottom=None, col_end=current_column + 1)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")

    start_row = current_row
    total_count1, grand_total_amount1 = 0, 0
    total_count2, grand_total_amount2 = 0, 0
    total_count1, grand_total_amount1, current_row = ppmp_item_category(office_supplies, ws, current_row)
    if lab_supplies:
        current_row += 1
        ws[f"{num_to_letter(current_column)}{current_row}"] = "LAB SUPPLIES"
        ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column + 1)}{current_row}")
        set_border_to_cell(ws, current_column, current_row, left=None, right=None, top=None, bottom=None,
                           col_end=current_column + 1)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
        total_count2, grand_total_amount2, current_row = ppmp_item_category(lab_supplies, ws, current_row)
    total_count = total_count1 + total_count2
    grand_total_amount = grand_total_amount1 + grand_total_amount2
    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "GRAND TOTAL:"
    ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column + 1)}{current_row}")
    set_border_to_cell(ws, current_column, current_row, col_end=current_column + 1)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center", underline="single")

    current_column = 16
    ws[f"{num_to_letter(current_column)}{current_row}"] = total_count
    set_border_to_cell(ws, current_column, current_row)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center",)

    current_column += 2
    ws[f"{num_to_letter(current_column)}{current_row}"] = grand_total_amount
    set_border_to_cell(ws, current_column, current_row)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center",)
    gray_fill = PatternFill(
        fill_type="solid",
        start_color="A6A6A6",
        end_color="A6A6A6"
    )

    for cell in ws[current_row]:
        cell.fill = gray_fill

    end_row = current_row

    current_column, current_row = set_signatories(ws, current_column, current_row)

    set_number_comma(ws, start_number_column, end_number_column, start_row, end_row)
    set_number_decimal(ws, start_decimal_column, end_decimal_column, start_row, end_row)

    current_column, current_row = set_dimensions(ws, current_column, current_row)

    set_border_to_cell(
        ws,
        col_start=start_column,
        row_start=start_row,
        col_end=end_column,
        row_end=end_row
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{title}.xlsx"'
    wb.save(response)
    return response



def num_to_letter(num):
    return chr(num + 64)

def letter_to_num(letter):
    return ord(letter.upper()) - 64


def set_format_to_cell(ws, column, row, font, size, bold, italic, horizontal, vertical, underline=None, wrap_text=False):
    ws[f"{num_to_letter(column)}{row}"].font = Font(bold=bold, italic=italic, underline=underline, name=font, size=size)
    ws[f"{num_to_letter(column)}{row}"].alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)

def set_border_to_cell(ws, col_start, row_start,
                       col_end=None, row_end=None,
                       left="thin", right="thin", top="thin", bottom="thin"):
    col_end = col_end or col_start
    row_end = row_end or row_start

    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            border = Border(
                left=Side(style=left) if c == col_start else Side(style="thin"),
                right=Side(style=right) if c == col_end else Side(style="thin"),
                top=Side(style=top) if r == row_start else Side(style="thin"),
                bottom=Side(style=bottom) if r == row_end else Side(style="thin"),
            )
            ws.cell(row=r, column=c).border = border

def set_number_comma(ws, start_column, end_column, start_row, end_row):
    for i in range(start_row, end_row + 1):
        for j in range(start_column, end_column + 1):
            ws[f"{num_to_letter(j)}{i}"].number_format = '#,##0'

def set_number_decimal(ws, start_column, end_column, start_row, end_row, decimal_number=2):
    for i in range(start_row, end_row + 1):
        for j in range(start_column, end_column + 1):
            ws[f"{num_to_letter(j)}{i}"].number_format = '#,##0.' + ("0" * decimal_number)

def set_dimensions(ws, current_column, current_row):
    current_column = 1
    ws.column_dimensions[num_to_letter(current_column)].width = 5
    current_column += 1
    ws.column_dimensions[num_to_letter(current_column)].width = 45
    current_column += 1
    ws.column_dimensions[num_to_letter(current_column)].width = 20
    current_column += 1
    for i in range(12):
        if i is 8:
            ws.column_dimensions[num_to_letter(current_column)].width = 13
        elif i > 8:
            ws.column_dimensions[num_to_letter(current_column)].width = 9
        else:
            ws.column_dimensions[num_to_letter(current_column)].width = 8
        current_column += 1

    ws.column_dimensions[num_to_letter(current_column)].width = 16
    current_column += 1
    ws.column_dimensions[num_to_letter(current_column)].width = 15
    current_column += 1
    ws.column_dimensions[num_to_letter(current_column)].width = 19
    current_column += 1
    return current_column, current_row

def set_header(ws, current_column, current_row, year):
    ws.merge_cells(f"A{current_row}:R{current_row}")
    ws[f"A{current_row}"] = "PROJECT PROCUREMENT MANAGEMENT PLAN (PPMP) " + year
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    ws.column_dimensions[num_to_letter(current_column)].width = 20
    current_row += 1
    ws[f"A{current_row}"] = "END-USER/UNIT: CICT"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, True, "left", "center")
    current_row += 1
    ws[f"A{current_row}"] = "Source of Fund: "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, True, "left", "center")

    ws.merge_cells("A5:A7")
    ws.merge_cells("B5:B7")
    ws.merge_cells("C5:C7")
    ws.merge_cells("D5:O5")
    ws.merge_cells("P5:P7")
    ws.merge_cells("Q5:Q7")
    ws.merge_cells("R5:R7")

    current_row = 5
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Seq."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "GENERAL DESCRIPTION"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Unit of Measure"
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Number of Units Needed"
    set_border_to_cell(ws, current_column, current_row, col_end=current_column + 11)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    current_column = 16
    ws[f"{num_to_letter(current_column)}{current_row}"] = "TOTAL"
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Price as per \nCatalogue"
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center",
                       wrap_text=True)
    start_decimal_column = current_column
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "TOTAL AMOUNT"
    set_border_to_cell(ws, current_column, current_row, row_end=current_row + 2)
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
    end_decimal_column = current_column
    end_number_column = current_column
    end_column = current_column

    current_row += 1
    current_column = 4
    start_number_column = current_column

    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October",
              "November", "December"]
    column_width = 8
    for i in range(0, len(months)):
        if i == 8:
            column_width = 13
        if i < 8:
            column_width = 9
        ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column)}{current_row + 1}")
        ws[f"{num_to_letter(current_column)}{current_row}"] = months[i]
        set_border_to_cell(ws, current_column, current_row, row_end=current_row + 1)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center")
        current_column += 1

    return (
        current_column,
        current_row,
        start_number_column,
        end_number_column,
        start_decimal_column,
        end_decimal_column,
        end_column,
    )

def set_signatories(ws, current_column, current_row):

    current_row += 2
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "NOTE: "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", underline="single")

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "1. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Provide all necessary information."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "2. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Several items may not be included in BulSU's consolidated catalogue. Thus, you may need provide your own description and estimated cost."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "3. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Items that are included in previous years PPMP that were not procured and that are needed in CY2024 may be included in this PPMP."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "4. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "The drop down list in column B may help guide the preparer of this PPMP to find the items needed by the colleges and offices. Nonetheless, the prepaper may manually search in the Catalogue Sheet of this PPMP form."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "5. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = 'Kindly delete all " #N/A " remarks once done to prevent summazation error.'
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_row += 1
    current_column = 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "6. "
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center",)
    current_column += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "For events, kindly include only the materials which will used."
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )
    current_column = 1


    current_row += 2
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Prepared by:"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 2
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Noted by:"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 5
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Recommending Approval:"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 4
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Approved by:"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    ppmp_signatories = private_supabase.table("DOCUMENT_SIGNATORY").select("*").eq("DocumentType", "PPMP DOCUMENT").execute()
    if ppmp_signatories is None:
        return None

    ppmp_signatories = ppmp_signatories.data

    current_row += 2
    current_column = 1

    end_user = [
        signatory["FullName"] for signatory in ppmp_signatories if signatory["PositionTitle"] == "End-user"
    ]
    if end_user:
        end_user = end_user[0]
    else:
        end_user = None
    ws[f"{num_to_letter(current_column)}{current_row}"] = end_user
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "left", "center", )

    current_column += 2

    budget_officer = [
        signatory["FullName"] for signatory in ppmp_signatories if signatory["PositionTitle"] == "Budget Officer"
    ]
    if budget_officer:
        budget_officer = budget_officer[0]
    else:
        budget_officer = None
    ws[f"{num_to_letter(current_column)}{current_row}"] = budget_officer
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "left", "center", )


    current_column += 5

    chancellor = [
        signatory["FullName"] for signatory in ppmp_signatories if signatory["PositionTitle"] == "Chancellor, Main Campus"
    ]
    if chancellor:
        chancellor = chancellor[0]
    else:
        chancellor = None
    ws[f"{num_to_letter(current_column)}{current_row}"] = chancellor
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "left", "center", )

    current_column += 4

    pres = [
        signatory["FullName"] for signatory in ppmp_signatories if signatory["PositionTitle"] == "University President"
    ]
    if pres:
        pres = pres[0]
    else:
        pres = None
    ws[f"{num_to_letter(current_column)}{current_row}"] = pres
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "left", "center", )

    current_column = 1
    current_row += 1
    ws[f"{num_to_letter(current_column)}{current_row}"] = "End-user"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 2
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Budget Officer"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 5
    ws[f"{num_to_letter(current_column)}{current_row}"] = "Chancellor, Main Campus"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    current_column += 4
    ws[f"{num_to_letter(current_column)}{current_row}"] = "University President"
    set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center", )

    return current_column, current_row

def ppmp_item_category(ppmp_category, ws, current_row):
    total_count = 0
    grand_total_amount = 0
    for item_category, items in ppmp_category.items():
        current_row += 1
        current_column = 1
        ws[f"{num_to_letter(current_column)}{current_row}"] = item_category
        ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column + 1)}{current_row}")
        set_border_to_cell(ws, current_column, current_row, col_end=current_column + 1)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, True, "center", "center")
        category_total_count = 0
        category_grand_total_amount = 0
        for ppmp_item in items:
            current_row += 1
            current_column = 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["Seq."]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")

            current_column += 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["GENERAL DESCRIPTION"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "left", "center")

            current_column += 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["Unit of Measure"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")

            current_column += 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["January"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")

            current_column += 1
            for i in range(11):
                ws[f"{num_to_letter(current_column)}{current_row}"] = ""
                set_border_to_cell(ws, current_column, current_row)
                set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")
                current_column += 1

            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["TOTAL"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")
            category_total_count += ppmp_item["TOTAL"]
            total_count += ppmp_item["TOTAL"]

            current_column += 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["Price as per Catalogue"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center", wrap_text=True)

            current_column += 1
            ws[f"{num_to_letter(current_column)}{current_row}"] = ppmp_item["TOTAL AMOUNT"]
            set_border_to_cell(ws, current_column, current_row)
            set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center")
            category_grand_total_amount += ppmp_item["TOTAL AMOUNT"]
            grand_total_amount += ppmp_item["TOTAL AMOUNT"]

        current_row += 1
        current_column = 1
        ws[f"{num_to_letter(current_column)}{current_row}"] = "Subtotal:"
        ws.merge_cells(f"{num_to_letter(current_column)}{current_row}:{num_to_letter(current_column + 1)}{current_row}")
        set_border_to_cell(ws, current_column, current_row, col_end=current_column + 1)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, True, "center", "center",)

        current_column += 3
        ws[f"{num_to_letter(current_column)}{current_row}"] = category_total_count
        set_border_to_cell(ws, current_column, current_row)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, False, False, "center", "center", )

        current_column = 16
        ws[f"{num_to_letter(current_column)}{current_row}"] = category_total_count
        set_border_to_cell(ws, current_column, current_row)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center", )

        current_column += 2
        ws[f"{num_to_letter(current_column)}{current_row}"] = category_grand_total_amount
        set_border_to_cell(ws, current_column, current_row)
        set_format_to_cell(ws, current_column, current_row, "Arial Narrow", 10, True, False, "center", "center", )

    return total_count, grand_total_amount, current_row
